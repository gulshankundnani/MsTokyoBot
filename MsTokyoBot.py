from telethon import TelegramClient, events, sync
from telethon import functions, types
from telethon.tl.functions.messages import (GetHistoryRequest)
from telethon.tl.functions.messages import (GetMessagesRequest)
from telethon.tl.functions.messages import (SendMediaRequest)
from telethon import errors
from PyDictionary import PyDictionary
from googletrans import Translator
from bs4 import BeautifulSoup
from PIL import Image
import openpyxl
from openpyxl import Workbook
import os
import sys
import time
import random
import pyjokes
import wikipedia
import googletrans
import json
import requests
import re
import urllib.request as urllib2



api_id = 1431692
api_hash = '4a91977a702732b8ba14fb92af6b1c2f'
happy_words = ['Cheerful','Contented','Delighted','Ecstatic','Elated','Joyous','Overjoyed','Pleased','Blissful','Chuffed','Delighted','Glad','Gratified','Joyful','Joyous','Pleased','Satisfied','Thankful','Tickled']
sad_words = ['Hopeless','Depressed','Mournful','Despairing','Miserable','Downcast','Gloomy','Heartbroken','Sorrowful','Glum','Dispirited','Dejected','Defeated','Woeful','Disheartened','Crushed','Crestfallen','Dismayed','Dismal','','Dreary']

commands = '++ [increase reputation (reply to a message/person)]\n\n'
commands += '-- [decrease reputation (reply to a message/person)]\n\n'
commands += 'twhat [get meaning of word (i.e. twhat happy)]\n\n'
commands += 'tjoke [get random joke]\n\n'
commands += 'tmeme [get random meme]\n\n'
commands += 'tlangcodes [get all language codes]\n\n'
commands += 'ttranslate [translate the terms(i.e. ttranslate [language code][term])]\n\n'
commands += 'tcommands [get all commands]\n\n'

client = TelegramClient('MsTokyoBot', api_id, api_hash)
client.start()


async def get_soup(url,header):
    return BeautifulSoup(urllib2.urlopen(urllib2.Request(url,headers=header)),'html.parser')

async def get_meaning(word):
    print(word)
    return wikipedia.summary(word, sentences = 1, auto_suggest=False, redirect=True)

async def checkFileExists(PATH):
    if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
        return 'yes'
    else:
        wb_obj = Workbook()
        sheet_obj = wb_obj.active
        wb_obj.save(filename=PATH)
        return 'yes'

async def getLatestReputation(username,filename):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active
    for row in range(1, sheet_obj.max_row+1):
        cell = sheet_obj.cell(row=row, column=1)
        if cell.value is not None and cell.value == username:
            cell2 = sheet_obj.cell(row=row, column=2)        
            break
    return cell2.value

async def editReputation(filename,toUsername,opr):
    try:        
        returnValue = toUsername
        check_file_exist = await checkFileExists(filename)
        if 'yes' in check_file_exist:
            if 'add' in opr:
                wb_obj = openpyxl.load_workbook(filename)
                sheet_obj = wb_obj.active
                for row in range(1, sheet_obj.max_row+1):
                    cell = sheet_obj.cell(row=row, column=1)
                    cell2 = sheet_obj.cell(row=row, column=2)
                    if cell.value is not None and cell.value == toUsername:
                        if cell2.value is not None:
                            cell2.value = (int(cell2.value) + 1)
                        wb_obj.save(filename=filename)
                    else:
                        cell.value = toUsername
                        cell2.value = 1
                        wb_obj.save(filename=filename)
                        
                        
            if 'sub' in opr:
                wb_obj = openpyxl.load_workbook(filename)
                sheet_obj = wb_obj.active
                for row in range(1, sheet_obj.max_row+1):
                    cell = sheet_obj.cell(row=row, column=1)
                    cell2 = sheet_obj.cell(row=row, column=2)
                    if cell.value is not None and cell.value == toUsername:
                        cell2.value = (int(cell2.value) - 1)
                        wb_obj.save(filename=filename)            

    except Exception as e:
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
    finally:
        return returnValue
    

async def eventData(event):
    replytomsgid = event.message.reply_to_msg_id
    if replytomsgid is not None:
        try:
            fromUserId = event.from_id
            fromUserEntity = await client.get_entity(fromUserId)
            fromUserName = fromUserEntity.username
            channelId = event.message.to_id.channel_id
            channelEntity = await client.get_entity(channelId)
            filename = './' + str(channelId) + '.xlsx'
            msgSearch = await client.get_messages(channelId, ids=replytomsgid)
            toUserEntity = await client.get_entity(msgSearch.from_id)
            toUserName = toUserEntity.username
            eventDataList = [fromUserId,fromUserEntity,fromUserName,channelId,channelEntity,filename,msgSearch,toUserEntity,toUserName]
            return eventDataList
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
    else:
        print("private chat")
        #await event.reply('Sorry for inconvinience! I am not available for personal chats. I only work with group chats.')    

@client.on(events.NewMessage)
async def my_event_handler(event):
    if 'tmeme' in event.raw_text.lower():
        try:
            channelId = event.message.to_id.channel_id
            channelEntity = await client.get_entity(channelId)
            query = "meme" # you can change the query for the image  here
            image_type="ActiOn"
            query= query.split()
            query='+'.join(query)
            url="https://www.google.com/search?q="+query+"&source=lnms&tbm=isch"
            print(url)
            header={'User-Agent':"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36"
            }
            soup = await get_soup(url,header)
            images = soup.findAll('img')
            im = random.choice(images)
            await event.reply( '.' + im['src'])
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)

    if 'tlangcodes' in event.raw_text.lower():
        try:
            s = googletrans.LANGUAGES
            listToStr = json.dumps(s)
            res = "\n".join("{!r}: {!r},".format(k, v) for k, v in s.items())
            await event.reply(res.rstrip(','))
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)

    if 'ttranslate' in event.raw_text.lower():
        try:
            cmd = event.raw_text.lower()
            to_lang = cmd.replace('ttranslate ','').split(' ')[0]
            term = cmd.replace('ttranslate ','').replace(to_lang + ' ','')
            print(term)
            translator= Translator()
            translation = translator.translate(term,dest=to_lang)
            await event.reply(translation.text)
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
    
    if '++' in event.raw_text.lower():
        try:
            eventDta = await eventData(event)
            if eventDta is None:
                print('')
            else:
                fromUserId = eventDta[0]
                fromUserEntity = eventDta[1]
                fromUserName = eventDta[2]
                channelId = eventDta[3]
                channelEntity = eventDta[4]
                filename = eventDta[5]
                msgSearch = eventDta[6]
                toUserEntity = eventDta[7]
                toUserName = eventDta[8]
                if fromUserName == toUserName:
                    await event.reply('You cannot give Karma to yourself !')
                else:
                    returnedToUserName = await editReputation(filename, toUserName, 'add')
                    latest_rep = await getLatestReputation(toUserName,filename)
                    await event.reply(random.choice(happy_words) + '! ' + '@' + fromUserName + ' increased reputation of @' + returnedToUserName + ' by 1 point. Total Reputation : ' + str(latest_rep))    
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
            
    if '--' in event.raw_text.lower():
        #print(event)
        try:
            eventDta = await eventData(event)
            if eventDta is None:
                print('')
            else:
                fromUserId = eventDta[0]
                fromUserEntity = eventDta[1]
                fromUserName = eventDta[2]
                channelId = eventDta[3]
                channelEntity = eventDta[4]
                filename = eventDta[5]
                msgSearch = eventDta[6]
                toUserEntity = eventDta[7]
                toUserName = eventDta[8]
                if fromUserName == toUserName:
                    await event.reply('You cannot give Karma to yourself !')
                else:
                    returnedToUserName = await editReputation(filename, toUserName, 'sub')
                    latest_rep = await getLatestReputation(toUserName,filename)
                    await event.reply(random.choice(sad_words) + '! ' + '@' + fromUserName + ' decreased reputation of @' + toUserName + ' by 1 point. Total Reputation : ' + str(latest_rep))    
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)

    if 'tjoke' in event.raw_text.lower():
        await event.reply(pyjokes.get_joke())

    if 'twhat' in event.raw_text.lower():
        try:            
            search = event.raw_text.lower().replace('twhat ','')
            dictionary = PyDictionary(search)
            info = await get_meaning(search)
            print(info)
            await event.reply(info)
        except Exception as e:
            if "DisambiguationError" in type(e).__name__:
                s = e.options
                listToStr = ' '.join([str(elem + '\n') for elem in s])
                await event.reply(search + ' could have following meanings. \n' + listToStr)
            else:
                await event.reply('Try another word.')
                print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)

    if 'tcommands' in event.raw_text.lower():
        try:
            await event.reply(commands)
        except Exception as e:
            print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
            
client.run_until_disconnected()
